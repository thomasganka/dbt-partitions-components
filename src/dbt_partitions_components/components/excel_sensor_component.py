"""Excel File Sensor Component.

Watches for changes to an Excel file (e.g., front office roster/salary data)
and triggers asset materialization when the file is modified. Demonstrates
event-driven pipelines where business users drop updated spreadsheets.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

import dagster as dg
from dagster.components import Component, ComponentLoadContext, Resolvable
from pydantic import BaseModel


class ExcelFileConfig(BaseModel):
    """Configuration for an Excel file to watch."""

    name: str
    file_path: str
    sheet_name: str = "Sheet1"
    description: str = ""
    group_name: str = "excel_ingestion"


def _generate_demo_excel(file_path: str, name: str) -> None:
    """Generate a sample NBA salary/roster Excel file for demo mode."""
    import random

    from openpyxl import Workbook

    random.seed(42)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    teams = [
        ("BOS", "Boston Celtics"),
        ("MIL", "Milwaukee Bucks"),
        ("DEN", "Denver Nuggets"),
        ("PHX", "Phoenix Suns"),
        ("LAL", "Los Angeles Lakers"),
        ("GSW", "Golden State Warriors"),
        ("DAL", "Dallas Mavericks"),
        ("MIA", "Miami Heat"),
    ]

    if name == "salary_cap":
        ws.append([
            "team_abbr", "team_name", "total_salary", "cap_space",
            "luxury_tax_bill", "num_players", "avg_salary", "cap_status",
        ])
        for abbr, full_name in teams:
            total = round(random.uniform(100_000_000, 190_000_000), 2)
            cap = 136_000_000
            space = cap - total
            tax = max(0, total - 165_000_000) * 1.5
            num_players = random.randint(13, 17)
            ws.append([
                abbr, full_name, total, space, round(tax, 2),
                num_players, round(total / num_players, 2),
                "Over Cap" if space < 0 else "Under Cap",
            ])
    elif name == "trade_targets":
        ws.append([
            "player_name", "team_abbr", "position", "age",
            "contract_value", "years_remaining", "trade_value_rating",
            "scouting_notes",
        ])
        players = [
            ("Marcus Johnson", "BOS", "SG", 26, 18_500_000, 2, 8.5, "Elite 3PT shooter"),
            ("DeAndre Williams", "MIL", "PF", 24, 12_000_000, 3, 7.8, "High upside two-way"),
            ("Tyler Chen", "DEN", "PG", 28, 25_000_000, 1, 9.1, "All-Star caliber"),
            ("James Carter", "PHX", "C", 30, 30_000_000, 2, 6.5, "Declining mobility"),
            ("Luis Rodriguez", "LAL", "SF", 23, 5_000_000, 4, 8.9, "Breakout candidate"),
            ("Kevin O'Brien", "GSW", "SG", 27, 22_000_000, 2, 7.2, "Inconsistent effort"),
            ("Andre Thompson", "DAL", "PF", 25, 15_000_000, 3, 8.0, "Switchable defender"),
            ("Chris Park", "MIA", "PG", 29, 20_000_000, 1, 7.5, "Expiring contract asset"),
        ]
        for row in players:
            ws.append(list(row))
    else:
        ws.append(["id", "name", "value"])
        for i in range(10):
            ws.append([i + 1, f"item_{i}", random.randint(1, 100)])

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)


def _make_excel_asset(
    config: ExcelFileConfig,
    demo_mode: bool,
    project_root: str,
) -> dg.AssetsDefinition:
    """Factory: create an asset that reads from an Excel file."""

    resolved_path = str(Path(project_root) / config.file_path)

    @dg.asset(
        key=dg.AssetKey(["excel", config.name]),
        description=config.description or f"Data loaded from Excel file: {config.name}",
        kinds={"excel", "duckdb"},
        group_name=config.group_name,
        metadata={
            "file_path": resolved_path,
            "sheet_name": config.sheet_name,
            "demo_mode": str(demo_mode).lower(),
        },
        tags={"source": "excel", "layer": "raw"},
    )
    def excel_asset(context: dg.AssetExecutionContext) -> dg.MaterializeResult:
        from openpyxl import load_workbook

        if demo_mode:
            context.log.info(f"[DEMO] Ensuring sample Excel file exists: {resolved_path}")
            if not Path(resolved_path).exists():
                _generate_demo_excel(resolved_path, config.name)

        if not Path(resolved_path).exists():
            raise FileNotFoundError(f"Excel file not found: {resolved_path}")

        wb = load_workbook(resolved_path, read_only=True)
        ws = wb[config.sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) for h in rows[0]] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        records = [dict(zip(headers, row)) for row in data_rows]

        wb.close()

        # Write to Parquet for downstream dbt consumption
        import json
        import os

        import duckdb

        out_dir = os.path.join(project_root, "data", "raw", config.name)
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"{config.name}.parquet")

        conn = duckdb.connect()
        conn.sql("CREATE TABLE _tmp AS SELECT * FROM read_xlsx(?, sheet_name := ?)", params=[resolved_path, config.sheet_name])
        conn.execute(f"COPY _tmp TO '{out_path}' (FORMAT PARQUET)")
        conn.close()

        context.log.info(f"Loaded {len(records)} rows from {resolved_path} -> {out_path}")

        return dg.MaterializeResult(
            metadata={
                "row_count": len(records),
                "columns": headers,
                "file_path": resolved_path,
                "output_path": out_path,
                "demo_mode": demo_mode,
            }
        )

    return excel_asset


def _make_file_sensor(
    config: ExcelFileConfig,
    project_root: str,
    demo_mode: bool,
    job,
) -> dg.SensorDefinition:
    """Factory: create a sensor that watches an Excel file for changes."""

    resolved_path = str(Path(project_root) / config.file_path)

    @dg.sensor(
        name=f"{config.name}_file_sensor",
        description=f"Watches for changes to {config.file_path} and triggers materialization",
        job=job,
        minimum_interval_seconds=30,
        default_status=dg.DefaultSensorStatus.RUNNING,
    )
    def file_sensor(context: dg.SensorEvaluationContext) -> dg.SensorResult:
        import os

        # In demo mode, ensure the sample file exists
        if demo_mode and not Path(resolved_path).exists():
            _generate_demo_excel(resolved_path, config.name)
            context.log.info(f"[DEMO] Generated sample Excel file: {resolved_path}")

        if not os.path.exists(resolved_path):
            context.log.info(f"File not found yet: {resolved_path}")
            return dg.SensorResult(run_requests=[], cursor=context.cursor)

        # Check file modification time
        mtime = str(os.path.getmtime(resolved_path))

        if context.cursor == mtime:
            context.log.info(f"No changes to {resolved_path}")
            return dg.SensorResult(run_requests=[], cursor=mtime)

        context.log.info(
            f"Detected change in {resolved_path} (mtime: {mtime})"
        )

        return dg.SensorResult(
            run_requests=[dg.RunRequest(run_key=f"{config.name}_{mtime}")],
            cursor=mtime,
        )

    return file_sensor


@dataclass
class ExcelSensorComponent(Component, Resolvable):
    """Watches Excel files for changes and triggers asset materialization.

    Creates a sensor that monitors an Excel file's modification time.
    When the file changes (e.g., a business user updates a spreadsheet),
    the sensor triggers a run to reload the data into the pipeline.

    This is ideal for scenarios where non-technical stakeholders maintain
    data in Excel and the pipeline should react to their updates.

    Attributes:
        files: List of Excel files to watch and ingest.
        demo_mode: If true, generates sample Excel files for local testing.
    """

    files: Sequence[ExcelFileConfig] = field(default_factory=list)
    demo_mode: bool = False

    def build_defs(self, context: ComponentLoadContext) -> dg.Definitions:
        project_root = str(Path(context.path).parents[3])

        assets = []
        sensors = []
        jobs = []

        for file_config in self.files:
            asset = _make_excel_asset(
                config=file_config,
                demo_mode=self.demo_mode,
                project_root=project_root,
            )
            assets.append(asset)

            job = dg.define_asset_job(
                name=f"{file_config.name}_refresh",
                selection=dg.AssetSelection.assets(
                    dg.AssetKey(["excel", file_config.name])
                ),
                description=f"Triggered by sensor when {file_config.file_path} changes",
            )
            jobs.append(job)

            sensor = _make_file_sensor(
                config=file_config,
                project_root=project_root,
                demo_mode=self.demo_mode,
                job=job,
            )
            sensors.append(sensor)

        return dg.Definitions(assets=assets, sensors=sensors, jobs=jobs)
